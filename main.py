import tkinter as tk
from stopWatchClass import Stopwatch
import time
from plyer import notification  # Import the notification functionality from plyer
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from ttkthemes import ThemedTk
import pygame
from tkinter import ttk

from selectedButton import *
import operator
import json
import copy
import sys
from tkinter import filedialog

from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2 import service_account

class PomodoroApp:
    """
    @brief Implements a Pomodoro Timer application with focus, contemplation, and reflection modes.

    This class provides the core functionality for a Pomodoro timer application, 
    allowing users to manage tasks, distractions, and reflections. It integrates with 
    Google Sheets or Excel for data logging and includes sound effects and a user-friendly interface.

    @details
    Modes:
    1. **Focus Mode**:
        - Designed for users to concentrate on a specific task for a set duration.
        - Features include:
            - Ability to toggle between distraction and problem-solving frames.
                - A problem-solving frame with MetaBox widgets for breaking down problems
                - Distraction frame, to allow users to log distractions to better focus while practicing
            - A countdown timer displaying the remaining focus time.
            - End focus and discard focus options to either transition to reflection or return to contemplation.

    2. **Contemplation Mode**:
        - Prepares the user for focus mode by setting tasks, goals, and timers.
        - Features include:
            - Entry fields for task description, goal definition, and timer settings.
            - Customizable focus and break durations.
            - A "Switch to Focus" button to validate input and start the focus session.
            - Error handling to ensure valid input before transitioning to focus mode.

    3. **Reflection Mode**:
        - Allows users to review their performance and record reflections after focus mode.
        - Features include:
            - Displays task, goal, and distraction counts during the session.
            - A break timer to track resting periods.
            - A text area for users to write reflections about their session.
            - A satisfaction level input (1-10) for session rating.
            - Options to save the reflection data into Google Sheets or Excel or discard changes.

    Additional Features:
    - **MetaBox Integration**: Provides a GUI tailored to allow users to break problems down to smaller digestable ones
    - **Sound Toggle**: Enables or disables ticking sounds for focus sessions.
    - **Distraction Management**: Counts distractions during focus mode.
    - **Data Integration**:
        - Authenticates with Google Sheets for data logging.
        - Falls back to Excel-based logging in case of API issues.

    @class PomodoroApp
    """
    def __init__(self, master):
        # self.reflection_textbox_visible = False
        # self.problem_solving_frame_visible = False  # Frame visibility state
        self.active_frame = None
        self.numOfDistraction = 0
        self.longDistraction = 0

        self.master = master
        pygame.init()
        pygame.mixer.init()  # Initialize the mixer module
        # self.tick_sound = pygame.mixer.Sound("dripdrop.mp3")  # Load your MP3 file
        self.tick_sound = self.load_sound_file()  # Dynamically load the MP3

        self.tick_channel = pygame.mixer.find_channel()
        self.sound_enabled = True  # Default to sound being on

        self.setup_frames()
        self.setup_focus_widgets()
        self.setup_contemplation_widgets()
        self.setup_reflection_widgets()
        self.show_contemplation()

    def setup_frames(self):
        style = ttk.Style()
        style.configure('TFrame', background='#383838')
        style.configure('TLabel', font=('Helvetica', 12), foreground='#E0E0E0', background='#383838')
        style.configure('TEntry', font=('Helvetica', 12), foreground='#D0D0D0', background='#505050', fieldbackground='#505050', borderwidth=1, relief='flat')
        style.configure('TButton', font=('Helvetica', 12), foreground='#E0E0E0', background='#404040', borderwidth=0, relief='flat')

        style.map('TButton',
                  foreground=[('pressed', '#FFFFFF'), ('active', '#FFFFFF')],
                  background=[('pressed', '#333333'), ('active', '#333333')],
                  relief=[('pressed', 'flat'), ('active', 'flat')])
        style.map('TEntry',
                  fieldbackground=[('focus', '#606060')],
                  foreground=[('focus', '#FFFFFF')])

        self.focus_frame = ttk.Frame(self.master, style='TFrame')
        self.contemplation_frame = ttk.Frame(self.master, style='TFrame')
        self.reflection_frame = ttk.Frame(self.master, style='TFrame')

        self.frames = [self.focus_frame, self.contemplation_frame, self.reflection_frame]

        for frame in self.frames:
            frame.pack(fill='both', expand=True)


        self.focus_button_frame = ttk.Frame(self.focus_frame, style='TFrame')

        self.focus_problemsolving_frame = ttk.Frame(self.master, style='TFrame')
        self.focus_distraction_frame = ttk.Frame(self.master, style='TFrame')


        self.focus_goalSetting_frame = ttk.Frame(self.master, style='TFrame')

        self.distractionTextBox = tk.Text(self.focus_distraction_frame, height=12, width=70)
        self.distractionTextBox.pack(pady=(5, 5))#, 

        self.meta_box_GoalSetting_app = MetaBoxApp(self.focus_goalSetting_frame)
        self.meta_box_app = MetaBoxApp(self.focus_problemsolving_frame)

        self.focus_frame_list = [self.focus_problemsolving_frame,self.focus_distraction_frame]

        for frame in self.focus_frame_list:
            frame.pack(fill='both', expand=True)

        self.focus_problemsolving_frame.pack_forget()
        self.focus_distraction_frame.pack_forget()

        self.focus_frame.pack_forget()
        self.reflection_frame.pack_forget()
        self.contemplation_frame.pack_forget()
        self.focus_goalSetting_frame.pack_forget()

    def setup_focus_widgets(self):
        self.focus_frame.config(style='TFrame')
        self.focus_frame.pack(fill='both', expand=True, pady=(10, 0), anchor='n')

        self.focus_task_label = ttk.Label(self.focus_frame, text="TASK", style='TLabel')
        self.focus_task_label.pack(pady=(10, 0), anchor='n')

        self.focus_goal_label = ttk.Label(self.focus_frame, text="Goal: Work on the project (1-2 sentences)", style='TLabel')
        self.focus_goal_label.pack(pady=(5, 10))

        self.timer_label = ttk.Label(
            self.focus_frame,
            text="25:00",
            font=('Helvetica', 24),
            foreground='#E0E0E0',
            background='#383838'
        )
        self.timer_label.pack(pady=(5, 20))


        self.focus_button_frame.pack(pady=(5, 5))

        # Single distraction button
        distraction_button = ttk.Button(
            self.focus_button_frame,
            text="DISTRACTION",
            style='TButton',
            command=lambda: self.toggle_focus_frames(frame_name="distraction")
        )
        distraction_button.pack(side='right', padx=(10, 20))

        problem_solving_button = ttk.Button(
            self.focus_button_frame,
            text="PROBLEM SOLVING",
            style='TButton',
            command=lambda: self.toggle_focus_frames(frame_name="problem_solving")
        )
        problem_solving_button.pack(side='left', padx=(5, 10))
        

        end_focus_button = ttk.Button(
            self.focus_frame,
            text="END FOCUS",
            style='TButton',
            command=self.show_reflection
        )
        end_focus_button.pack(pady=(20, 10))

        discard_focus_button = ttk.Button(
            self.focus_frame,
            text="DISCARD",
            style='TButton',
            command=self.show_contemplation
        )
        discard_focus_button.pack(pady=(20, 10))

        self.toggle_sound_button = ttk.Button(
            self.focus_frame,
            text="TOGGLE SOUND",
            style='TButton',
            command=self.toggle_sound
        )
        self.toggle_sound_button.pack(pady=(10, 0))

        self.pomodoro = Stopwatch(self.timer_label)
    def setup_contemplation_widgets(self):

        self.task_label = ttk.Label(self.contemplation_frame, text="TASK", style='TLabel')
        self.task_label.pack(pady=(10, 2), anchor='w')

        self.task_entry = ttk.Entry(self.contemplation_frame, style='TEntry', width=40)
        self.task_entry.pack(pady=(2, 10))
        self.task_entry.bind("<Tab>", lambda e: self.goal_entry.focus_set())

        self.goal_label = ttk.Label(self.contemplation_frame, text="GOAL", style='TLabel')
        self.goal_label.pack(pady=(10, 2), anchor='w')
        self.goal_entry = ttk.Entry(self.contemplation_frame, style='TEntry', width=60)
        self.goal_entry.pack(pady=(2, 10))

        timer_frame = ttk.Frame(self.contemplation_frame, style='TFrame')
        timer_frame.pack(pady=(10, 10))

        pomodoro_label = ttk.Label(timer_frame, text="TIMER(MIN):", style='TLabel')
        pomodoro_label.pack(side='left', padx=(5, 2))
        self.pomodoro_entry = ttk.Entry(timer_frame, style='TEntry', width=5)
        self.pomodoro_entry.pack(side='left', padx=(2, 20))
        self.pomodoro_entry.insert(0, '25')

        break_label = ttk.Label(timer_frame, text="REST TIMER(MIN):", style='TLabel')
        break_label.pack(side='left', padx=(5, 2))
        self.break_entry = ttk.Entry(timer_frame, style='TEntry', width=5)
        self.break_entry.pack(side='left')
        self.break_entry.insert(0, '5')

        contemplation_button = ttk.Button(self.contemplation_frame, text="SWITCH TO FOCUS", style='TButton', command=self.show_focus)
        contemplation_button.pack(pady=(10, 10))

        self.contemplation_frame.pack(fill='both', expand=True, padx=10, pady=10)
    def setup_reflection_widgets(self):
        self.reflection_frame.config(style='TFrame')
        self.reflection_frame.pack(fill='both', expand=True)

        self.reflection_task_label = ttk.Label(
            self.reflection_frame,
            text="TASK",
            style='TLabel'
        )
        self.reflection_task_label.pack(pady=(10, 0), anchor='n')

        self.reflection_goal_label = ttk.Label(
            self.reflection_frame,
            text="Goal: Work on the project (1-2 sentences)",
            style='TLabel'
        )
        self.reflection_goal_label.pack(pady=(5, 10))

        # Single distraction display
        distraction_frame = ttk.Frame(self.reflection_frame, style='TFrame')
        distraction_frame.pack(pady=(5, 10))

        distraction_label = ttk.Label(distraction_frame, text="Distractions:", style='TLabel')
        self.distraction_value = ttk.Label(distraction_frame, text=str(self.numOfDistraction), style='TLabel')

        distraction_label.pack(side='left')
        self.distraction_value.pack(side='left')

        self.break_time_label = ttk.Label(
            self.reflection_frame,
            text="00:00",
            style='TLabel',
            font=('Helvetica', 24)
        )
        self.break_time_label.pack(pady=(5, 10))

        self.reflection_textbox = tk.Text(self.reflection_frame, height=12, width=70)
        self.reflection_textbox.pack(pady=(5, 5))

        satisfaction_label = ttk.Label(
            self.reflection_frame,
            text="SATISFACTION LEVEL (1-10):",
            style='TLabel'
        )
        satisfaction_label.pack(pady=(5, 2))

        self.satisfaction_entry = ttk.Entry(self.reflection_frame, style='TEntry', width=5)
        self.satisfaction_entry.pack(pady=(2, 10))
        self.satisfaction_entry.bind("<FocusOut>", self.validate_satisfaction)

        button_frame = ttk.Frame(self.reflection_frame, style='TFrame')
        button_frame.pack(pady=(10, 10))

        insert_button = ttk.Button(
            button_frame,
            text="INSERT",
            style='TButton',
            command=lambda: self.show_contemplation("insert")
        )
        discard_button = ttk.Button(
            button_frame,
            text="DISCARD",
            style='TButton',
            command=lambda: self.show_contemplation("discard")
        )
        insert_button.pack(side='left', padx=10)
        discard_button.pack(side='left')

        self.breakStopwatch = Stopwatch(self.break_time_label)
        
    def toggle_goalSetting_frame(self):
        """
        Toggles visibility of the goal-setting frame (focus_goalSetting_frame) as a fullscreen overlay.
        """
        if self.active_frame == self.focus_goalSetting_frame:
            # If the goal-setting frame is active, hide it
            self.focus_goalSetting_frame.place_forget()
            self.active_frame = None
        else:
            # Show the goal-setting frame as a fullscreen overlay
            self.focus_goalSetting_frame.place(relx=0, rely=0, relwidth=1, relheight=1)
            self.focus_goalSetting_frame.tkraise()  # Bring this frame to the top
            self.active_frame = self.focus_goalSetting_frame


    def show_focus(self):
        error_present = False

        # Reset field text color
        self.task_entry.config(foreground='white')
        self.pomodoro_entry.config(foreground='white')
        self.break_entry.config(foreground='white')

        # Validate task entry
        if not self.task_entry.get().strip():
            self.task_entry.delete(0, 'end')
            self.task_entry.insert(0, "Error: Task cannot be empty.")
            self.task_entry.config(foreground='red')
            error_present = True

        # Validate pomodoro and break entries
        try:
            pomodoro_time = int(self.pomodoro_entry.get())
            break_time = int(self.break_entry.get())
            if pomodoro_time <= 0 or break_time <= 0:
                raise ValueError("Time must be positive integers.")
        except ValueError:
            self.pomodoro_entry.delete(0, 'end')
            self.pomodoro_entry.insert(0, "Error")
            self.pomodoro_entry.config(foreground='red')
            self.break_entry.delete(0, 'end')
            self.break_entry.insert(0, "Error")
            self.break_entry.config(foreground='red')
            error_present = True

        # If no errors, proceed to focus mode
        if not error_present:
            self.toggle_sound(state="ON")
            task_text = self.task_entry.get()
            goal_text = self.goal_entry.get()
            self.focus_task_label.config(text="Task: " + task_text)
            self.focus_goal_label.config(text="Goal: " + goal_text)

            # Reset and start pomodoro timer
            self.pomodoro.running = False
            self.pomodoro.in_bonus = False
            self.pomodoro.set_time(pomodoro_time)
            self.pomodoro.start()

            # Switch frames
            self.contemplation_frame.pack_forget()
            self.reflection_frame.pack_forget()
            self.focus_frame.pack(fill='both', expand=True)
        # print(self.meta_box_app.head)
        # if(self.meta_box_app.head is None):
            self.meta_box_app.new_head()

        self.master.update_idletasks()  # Refresh geometry calculations
        self.master.geometry("")  # Let Tkinter automatically resize the window
            
    def show_reflection(self):
        # Pause pomodoro and sound
        self.tick_channel.stop()
        self.pomodoro.pause()

        # Print data for debugging
        data = self.pomodoro.getStopWatchData()
        print(data)
        print("\nDistraction count: " + str(self.numOfDistraction))

        # Update reflection labels
        task_text = self.task_entry.get()
        goal_text = self.goal_entry.get()
        self.reflection_task_label.config(text="Task: " + task_text)
        self.reflection_goal_label.config(text="Goal: " + goal_text)
        self.distraction_value.config(text=str(self.numOfDistraction))

        distraction_text = self.distractionTextBox.get("1.0", "end-1c")  # Get text from distractionTextBox
        self.reflection_textbox.delete("1.0", "end")  # Clear the reflection_textbox
        self.reflection_textbox.insert("1.0", distraction_text)  # Insert text into reflection_textbox
        self.distractionTextBox.delete("1.0", "end")  # Clear the distractionTextBox

        # Reset and start break timer
        self.breakStopwatch.running = False
        self.breakStopwatch.in_bonus = False
        self.breakStopwatch.set_time(int(self.break_entry.get()))
        self.breakStopwatch.start()

        # Switch frames
        self.focus_frame.pack_forget()
        self.reflection_frame.pack_forget()
        self.contemplation_frame.pack_forget()
        self.reflection_frame.pack(fill='both', expand=True)
        self.master.update_idletasks()  # Refresh geometry calculations
        self.master.geometry("")  # Let Tkinter automatically resize the window

    def show_contemplation(self, action=None):
        error_present = False
        if action == "insert":
            if self.validate_data():
                self.insert_into_database()
            else:
                send_notification("error", "Data validation failed. Not inserting.")
                error_present = True
        elif action == "discard":
            print("Changes discarded.")

        if not error_present:
            self.toggle_sound(state="OFF")
            self.clear_form_fields()
            self.focus_frame.pack_forget()
            self.reflection_frame.pack_forget()
            self.contemplation_frame.pack_forget()
            self.contemplation_frame.pack(fill='both', expand=True)
            self.master.update_idletasks()  # Refresh geometry calculations
            self.master.geometry("")  # Let Tkinter automatically resize the window
    def validate_satisfaction(self, event=None):
        try:
            val = int(self.satisfaction_entry.get())
            if not 1 <= val <= 10:
                raise ValueError("Satisfaction rating must be between 1 and 10.")
            else:
                return True
        except ValueError:
            self.satisfaction_entry.delete(0, 'end')
            self.satisfaction_entry.insert(0, "Invalid input")
        return False
    def validate_data(self):
        task = self.task_entry.get().strip()
        if not task:
            send_notification("Error", "Task cannot be empty.")
            return False

        time_spent = self.pomodoro.get_time()
        if time_spent <= 0:
            send_notification("Error", "Time spent on task must be greater than zero.")
            return False

        bonus_time_spent = self.pomodoro.bonus_time
        if bonus_time_spent < 0:
            send_notification("Error", "Bonus time cannot be negative.")
            return False

        time_rested = self.breakStopwatch.get_time()
        if time_rested <= 0:
            send_notification("Error", "Time rested must be greater than zero.")
            return False

        bonus_time_rested = self.breakStopwatch.bonus_time
        if bonus_time_rested < 0:
            send_notification("Error", "Bonus time for rest cannot be negative.")
            return False

        short_distractions = self.numOfDistraction
        if short_distractions < 0:
            send_notification("Error", "Number of short-term distractions cannot be negative.")
            return False

        long_distractions = self.longDistraction
        if long_distractions < 0:
            send_notification("Error", "Number of long-term distractions cannot be negative.")
            return False

        return True

    def authenticate_google_sheets(self):
        """Authenticate and return the Google Sheets service."""
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        SERVICE_ACCOUNT_FILE = r'C:\Users\Parry\Documents\python\deliberatepra-9ecb1d8fcae6.json'

        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)

        try:
            service = build('sheets', 'v4', credentials=creds)
            return service
        except HttpError as err:
            print(f"An error occurred: {err}")
            return None
    def insert_into_database(self):
        SPREADSHEET_ID = '1ETycSikjSfLCkZQLMqMa2GOTWE3Z4peGKZnI7o60VzQ'
        RANGE_NAME = 'Sheet1!A1'  # Update this range based on your sheet structure

        # Pause timers
        self.pomodoro.pause()
        self.breakStopwatch.pause()

        # Prepare data
        data = {
            'Start Time': time.strftime("%Y-%m-%d %H:%M:%S", self.pomodoro.startLocalTime) if self.pomodoro.startLocalTime else "Not started",
            'Task': self.task_entry.get(),
            'Goal': self.goal_entry.get(),
            'Time Spent on Task': self.pomodoro.getStopWatchData()['Elapsed Time'],
            'Bonus Time Spent on Task': self.pomodoro.getStopWatchData()['Bonus Time'],
            'Time Rested': self.breakStopwatch.getStopWatchData()['Elapsed Time'],
            'Bonus Time Rested': self.breakStopwatch.getStopWatchData()['Bonus Time'],
            'Number of Distractions': self.numOfDistraction,
            'Reflection': self.reflection_textbox.get("1.0", "end-1c"),
        }

        # Convert data to a list of lists (for Google Sheets API)
        values = [[v for v in data.values()]]

        # Authenticate and get Google Sheets service
        service = self.authenticate_google_sheets()
        if service is None:
            print("Failed to authenticate Google Sheets API.")
            return

        # Append data to Google Sheets
        body = {'values': values}
        try:
            result = service.spreadsheets().values().append(
                spreadsheetId=SPREADSHEET_ID,
                range=RANGE_NAME,
                valueInputOption="RAW",  # Use 'RAW' to insert data as-is
                insertDataOption="INSERT_ROWS",  # Ensure new rows are added
                body=body
            ).execute()
            print(f"{result.get('updates').get('updatedCells')} cells appended.")
        except Exception as err:
            print(f"An error occurred with Google Sheets API: {err}. Falling back to Excel.")
            self.insert_into_database_excel()
    #case for when there is an issue connecting with GOOGLE API. Insert into excel sheet instead
    def insert_into_database_excel(self):
        filename = 'database.xlsx'
        self.pomodoro.pause()
        self.breakStopwatch.pause()
        data = self.pomodoro.getStopWatchData()
        restingData = self.breakStopwatch.getStopWatchData()
        start_time_formatted = time.strftime("%Y-%m-%d %H:%M:%S", self.pomodoro.startLocalTime) if self.pomodoro.startLocalTime else "Not started"
        data = {
            'Start Time': data['Local Time'],
            'Task': self.task_entry.get(),
            'Goal': self.goal_entry.get(),
            'Time Spent on Task': data['Elapsed Time'],
            'Bonus Time Spent on Task': data['Bonus Time'],
            'Time Rested':  restingData['Elapsed Time'],
            'Bonus Time Rested': restingData['Bonus Time'],
            'Number of ': self.numOfDistraction,
            'Reflection': self.reflection_textbox.get("1.0", "end-1c"),
            'Satisfaction Level': self.satisfaction_entry.get()
        }

        file_exists = os.path.isfile(filename)

        if not file_exists:
            wb = Workbook()
            ws = wb.active
        else:
            wb = load_workbook(filename)
            ws = wb.active

        if not file_exists:
            for col, header in enumerate(data.keys(), start=1):
                ws[f'{get_column_letter(col)}1'] = header

        next_row = ws.max_row + 1 if file_exists else 2

        for col, (key, value) in enumerate(data.items(), start=1):
            ws[f'{get_column_letter(col)}{next_row}'] = value

        wb.save(filename)
        print("Data inserted successfully into Excel.")

    def clear_form_fields(self):
        #delete problemSolving METAboxes
        

        # self.toggle_reflection_textbox(True)
        self.active_frame = None
        # Clear text in reflection labels
        if hasattr(self, 'task_entry'):
            self.reflection_task_label.config(text=" ")
        if hasattr(self, 'goal_entry'):
            self.reflection_goal_label.config(text=" ")
            
        # Clear satisfaction entry
        if hasattr(self, 'satisfaction_entry'):
            self.satisfaction_entry.delete(0, 'end')
            
        # Clear reflection_textbox
        if hasattr(self, 'reflection_textbox'):
            self.reflection_textbox.delete("1.0", "end")
            
        if hasattr(self, 'distractionTextBox'):
            self.distractionTextBox.delete("1.0", "end")
            
        # Reset breakStopwatch
        self.reset_and_pause_break_stopwatch()
        
        # Reset single distraction count
        if hasattr(self, 'distraction_value'):
            self.numOfDistraction = 0
            self.distraction_value.config(text=str(self.numOfDistraction))
            
        # Reset break time label
        if hasattr(self, 'break_time_label'):
            self.break_time_label.config(text="00:00")
        # self.meta_box_app.delete_tree(self.meta_box_app.head)

    def reset_and_pause_break_stopwatch(self):
        
        if self.breakStopwatch.running:
            self.breakStopwatch.pause()
        self.breakStopwatch.remaining_time = 0
        self.breakStopwatch.in_bonus = False
        self.breakStopwatch.bonus_time = 0
    def toggle_focus_frames(self, event=None, frame_name=None):
        """
        Toggles between the distraction and problem-solving frames.
        Args:
            event: Optional, triggered by a button press or binding.
            frame_name: The name of the frame to toggle ('distraction' or 'problem_solving').
        """
        # Define the frames dictionary
        frames = {
            "distraction": self.focus_distraction_frame,
            "problem_solving": self.focus_problemsolving_frame,
        }

        # Determine the frame to toggle
        requested_frame = frames.get(frame_name)

        if frame_name == "distraction":
            # Increment the distraction counter when toggling the distraction frame ON
            if frame_name == "distraction" and self.active_frame != requested_frame:
                self.numOfDistraction += 1

        # If no frame is active, or the same frame is requested, toggle visibility
        if self.active_frame is None:
            requested_frame.pack(fill="both", expand=True, after=self.focus_button_frame)
            self.active_frame = requested_frame
        elif self.active_frame == requested_frame:
            requested_frame.pack_forget()
            self.active_frame = None
        else:
            # If a different frame is active, hide it and show the requested one
            self.active_frame.pack_forget()
            requested_frame.pack(fill="both", expand=True, after=self.focus_button_frame)
            self.active_frame = requested_frame

        # Adjust the window size to fit the updated layout
        self.master.update_idletasks()  # Refresh geometry calculations
        self.master.geometry("")  # Let Tkinter automatically resize the window

    def load_sound_file(self):
        """
        Dynamically loads the MP3 file specified in the first row of configSound.txt.
        Falls back to default or raises an error if files are missing.
        """
        base_dir = os.path.dirname(os.path.abspath(__file__))
        config_file = os.path.join(base_dir, "configSound.txt")
        default_mp3 = os.path.join(base_dir, "dripdrop.mp3")

        # Check for configSound.txt
        if os.path.exists(config_file):
            with open(config_file, "r") as f:
                mp3_file = f.readline().strip()
                mp3_path = os.path.join(base_dir, mp3_file)
                if os.path.exists(mp3_path):
                    return pygame.mixer.Sound(mp3_path)
                else:
                    print(f"MP3 file {mp3_file} not found in {base_dir}.")
        else:
            print(f"{config_file} not found in {base_dir}.")

        # Fall back to default
        if os.path.exists(default_mp3):
            return pygame.mixer.Sound(default_mp3)
        else:
            raise FileNotFoundError(f"Default MP3 file missing: {default_mp3}")

        
    def toggle_sound(self, state=" "):
        
        if state == " ":
            self.sound_enabled = not self.sound_enabled
            self.toggle_sound_button.config(text="DISABLE SOUND" if self.sound_enabled else "ENABLE SOUND")
        elif state == "ON":
            self.sound_enabled = True
        elif state == "OFF":
            self.sound_enabled = False

        print("sound is " + str(self.sound_enabled))
        if self.sound_enabled:
            if not self.tick_channel.get_busy():
                self.tick_channel.play(self.tick_sound, loops=-1)
        else:
            self.tick_channel.stop()

    def save_active_meta_box(self):
        """
        Saves and opens the meta box app for the currently active frame.
        If the Goal Setting frame is active, it saves the Goal Setting meta box.
        Otherwise, it saves the Problem Solving meta box.
        """
        if self.active_frame == self.focus_goalSetting_frame:
            # Save and open for Goal Setting
            self.meta_box_GoalSetting_app.save(self.meta_box_GoalSetting_app.head)
        else:
            # Save and open for Problem Solving
            self.meta_box_app.save(self.meta_box_app.head)
    def open_active_meta_box(self):
        """
        Opens the meta box app for the currently active frame.
        If the Goal Setting frame is active, it opens the Goal Setting meta box.
        Otherwise, it opens the Problem Solving meta box.
        """
        if self.active_frame == self.focus_goalSetting_frame:
            # Open for Goal Setting
            self.meta_box_GoalSetting_app.open_file()
        else:
            # Open for Problem Solving
            self.meta_box_app.open_file()


def format_time(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"
def send_notification(title, message):
    notification.notify(
        title=title,
        message=message,
        app_name="Pomodoro Timer"
    )

def initializeBorderButtons(master,frame,_appInstance):


    global app
    menu = tk.Menu(master)
    master.config(menu = menu) 

    subMenu = tk.Menu(menu,tearoff = 0) # creating a menu item within the menu
    transparentMenu = tk.Menu(menu,tearoff = 0)

    menu.add_cascade(label = "Edit", menu = subMenu) #Name of drop down menu
    menu.add_cascade(label = "Transparancy",  menu = transparentMenu)
    subMenu.add_command(label = "save", command=lambda: app.save_active_meta_box())
    subMenu.add_command(label="Open",command=lambda: app.open_active_meta_box())
    subMenu.add_command(label="Goal setting" , command=app.toggle_goalSetting_frame)
    subMenu.add_separator()

    transparentMenu.add_command(label = "30%",command = lambda: changeRootTransparency(master,0.3))
    transparentMenu.add_command(label = "50%",command =lambda: changeRootTransparency(master,0.5) )
    transparentMenu.add_command(label = "60%",command = lambda: changeRootTransparency(master,0.6))
    transparentMenu.add_command(label = "70%",command = lambda: changeRootTransparency(master,0.7))
    transparentMenu.add_command(label = "80%",command = lambda: changeRootTransparency(master,0.8))
    transparentMenu.add_command(label = "100%",command = lambda: changeRootTransparency(master,1.0))	
def changeRootTransparency(master,percentage):
    root.attributes("-alpha",percentage)

def onFrameConfigure(canvas):
    '''Reset the scroll region to encompass the inner frame'''
    canvas.configure(scrollregion=canvas.bbox("all"))

if __name__ == "__main__":
   
    root = ThemedTk(theme="equilux")
    root.title("Drip Dropper")
    root.configure(background='#383838')

    app = PomodoroApp(root)
    
    mainCanvas = tk.Canvas(root, background = 'black')
    frame = tk.Frame(mainCanvas, background="black")
    framelst = []
    mainCanvas.configure(scrollregion=mainCanvas.bbox("all"))
    root.attributes("-alpha",1.0)

    initializeBorderButtons(root,app.focus_problemsolving_frame,_appInstance = app)

    root.geometry('400x250')
    root.mainloop()